"""
Genera Gantt Excel para el Plan de Acción – Integración SAP en Fabric
Ejecutar: python3 generar_gantt.py
Salida: Plan_Accion_SAP_Fabric_Gantt.xlsx
"""

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from datetime import date, timedelta
import copy

# ─────────────────────────────────────────────
# 1. DATOS DEL PLAN
# ─────────────────────────────────────────────

START_DATE = date(2026, 3, 17)   # Hoy

# Colores por fase (hex sin #)
PHASE_COLORS = {
    0: "4472C4",   # azul
    1: "70AD47",   # verde
    2: "ED7D31",   # naranja
    3: "9E48B5",   # violeta
    4: "FF0000",   # rojo (go-live crítico)
    5: "00B0F0",   # celeste
}

PHASE_LIGHT = {
    0: "D9E1F2",
    1: "E2EFDA",
    2: "FCE4D6",
    3: "EAD8F5",
    4: "FFD0D0",
    5: "DDEEFF",
}

PHASES = [
    {
        "num": 0,
        "name": "Cierre estrategia",
        "objetivo": "Estrategia adoptada, lista de entidades y plan source_system",
        "hito": "Priorización lista",
        "duration_days": 7,   # 1 semana (mínimo)
        "activities": [
            {
                "code": "0.1",
                "name": "Formalizar adopción Estrategia 3",
                "desc": "Validar con stakeholders Estrategia 3 (Silver dual + conformada); documentar ajustes.",
                "deliverable": "OK o ajustes al plan documentados",
                "dur": 3,
            },
            {
                "code": "0.2",
                "name": "Definir lista de entidades Silver",
                "desc": "Identificar y priorizar entidades Silver que alimentan vistas Gold para SEL/SCO/STU. Maestros primero.",
                "deliverable": "Lista priorizada (clientes, materiales, órdenes venta, facturas…)",
                "dur": 3,
            },
            {
                "code": "0.3",
                "name": "Identificar tablas sin source_system",
                "desc": "Revisar qué tablas Silver tienen source_system; plan de extensión sin romper pipelines QAD.",
                "deliverable": "Lista de tablas + plan de cambio faseado",
                "dur": 2,
            },
        ],
    },
    {
        "num": 1,
        "name": "Mapeo y diseño",
        "objetivo": "Matriz SAP→Silver por entidad, diseño lh_silver_sap y capa conformada, Bronze SAP ampliado",
        "hito": "Diseño y matriz listos",
        "duration_days": 15,  # 3 semanas
        "activities": [
            {
                "code": "1.1",
                "name": "Matriz de mapeo CDS ↔ Silver",
                "desc": "Por entidad: identificar CDS SAP (SE11/HANA Studio), mapeo campo a campo, tipos y reglas.",
                "deliverable": "Matriz de mapeo por entidad",
                "dur": 7,
            },
            {
                "code": "1.2",
                "name": "Full vs. Incremental + control SAP",
                "desc": "Definir Full/Incremental por CDS; registrar en source_to_bronze_control_sap.",
                "deliverable": "source_to_bronze_control_sap actualizado",
                "dur": 4,
            },
            {
                "code": "1.3",
                "name": "Diseño lh_silver_sap",
                "desc": "Definir tablas y columnas de lh_silver_sap con mismo contrato que Silver QAD.",
                "deliverable": "Documento de diseño lh_silver_sap",
                "dur": 3,
            },
            {
                "code": "1.4",
                "name": "Diseño capa conformada",
                "desc": "Definir vistas UNION por entidad (lh_silver_qad + lh_silver_sap); vistas vs tablas materializadas.",
                "deliverable": "Especificación capa conformada",
                "dur": 2,
            },
            {
                "code": "1.5",
                "name": "Ampliar lh_bronze_sap (piloto)",
                "desc": "Incluir CDS del piloto en lh_bronze_sap; registrar en source_to_bronze_control_sap.",
                "deliverable": "Pipelines Source→Bronze SAP para piloto",
                "dur": 3,
            },
        ],
    },
    {
        "num": 2,
        "name": "Piloto (1 entidad)",
        "objetivo": "Una entidad de punta a punta; patrón documentado para rollout",
        "hito": "Patrón replicable",
        "duration_days": 7,  # 1.5 semanas
        "activities": [
            {
                "code": "2.1",
                "name": "ETL Bronze SAP → lh_silver_sap",
                "desc": "Notebook/pipeline parametrizado para entidad piloto: mapeo, source_system, company_code.",
                "deliverable": "Pipeline + tabla en lh_silver_sap",
                "dur": 4,
            },
            {
                "code": "2.2",
                "name": "Vista conformada piloto",
                "desc": "UNION de Silver QAD y lh_silver_sap. Probar con filtro company_code / source_system.",
                "deliverable": "Vista conformada probada",
                "dur": 2,
            },
            {
                "code": "2.3",
                "name": "Validación punta a punta",
                "desc": "Validar volúmenes y muestreo SAP vs QAD; verificar reporte Gold/RPA con datos unificados.",
                "deliverable": "OK validación + lecciones aprendidas",
                "dur": 2,
            },
            {
                "code": "2.4",
                "name": "Documentar patrón estándar",
                "desc": "Guía reproducible: pasos, nombres, tablas de control para agregar una entidad SAP.",
                "deliverable": "Guía 'cómo agregar entidad SAP'",
                "dur": 1,
            },
        ],
    },
    {
        "num": 3,
        "name": "Rollout entidades Empresas TECC",
        "objetivo": "Todas las entidades necesarias para vistas Gold (reportes y RPAs) de SEL/SCO/STU",
        "hito": "Vistas listas para SEL",
        "duration_days": 35,  # 5 semanas
        "activities": [
            {
                "code": "3.1",
                "name": "Priorizar orden de entidades",
                "desc": "Maestros primero (clientes, materiales, proveedores), luego transaccionales; considerar dependencias.",
                "deliverable": "Orden de implementación documentado",
                "dur": 2,
            },
            {
                "code": "3.2",
                "name": "Implementar entidades (Bronze → Silver → Conformada)",
                "desc": "Por entidad: ampliar Bronze SAP, ETL→lh_silver_sap, vista conformada, registro control, prueba básica.",
                "deliverable": "Todas las entidades críticas en lh_silver_sap y capa conformada",
                "dur": 30,
            },
            {
                "code": "3.3",
                "name": "Revisión de calidad y estándares",
                "desc": "Revisar por lote: nombres, tipos, source_system, company_code; aplicar estándares.",
                "deliverable": "Ajustes y estándares documentados",
                "dur": 5,   # continuo, se solapa con 3.2
            },
            {
                "code": "3.4",
                "name": "Extender source_system en Silver",
                "desc": "Añadir source_system en tablas Silver que recibirán datos SAP y aún no lo tienen.",
                "deliverable": "Silver multi-origen habilitado",
                "dur": 2,
            },
        ],
    },
    {
        "num": 4,
        "name": "Go-live SEL",
        "objetivo": "Reportes y RPAs de SEL operativos con datos SAP vía capa conformada",
        "hito": "SEL en producción con SAP",
        "duration_days": 7,  # 1 semana
        "activities": [
            {
                "code": "4.1",
                "name": "Actualizar go-live y validación rápida",
                "desc": "Actualizar parámetros go-live con fecha real SEL; validar vistas conformadas y reportes.",
                "deliverable": "Tabla go-live actualizada; validación ejecutada",
                "dur": 1,
            },
            {
                "code": "4.2",
                "name": "Validación con negocio",
                "desc": "Validar reportes Gold y vistas RPA de SEL; comparar con reportes estándar SAP.",
                "deliverable": "OK de negocio o lista de ajustes",
                "dur": 4,
            },
            {
                "code": "4.3",
                "name": "Soporte post go-live SEL",
                "desc": "Atender incidencias de datos y ajustes menores de mapeo o filtros.",
                "deliverable": "Estabilización operativa SEL",
                "dur": 3,
            },
        ],
    },
    {
        "num": 5,
        "name": "STU + SCO + Bloques financieras",
        "objetivo": "Go-live STU (May) y SCO (Jun); estabilización final del proceso",
        "hito": "Transición cerrada",
        "duration_days": 25,  # 3.5 semanas
        "activities": [
            {
                "code": "5.1",
                "name": "Go-live STU + 2.º bloque financieras (Mayo)",
                "desc": "Actualizar go-live. Validación rápida por empresa reportes Gold/RPA; soporte inicial.",
                "deliverable": "STU + bloque fin. cubiertos en capa conformada",
                "dur": 7,
            },
            {
                "code": "5.2",
                "name": "Go-live SCO + último bloque financieras (Junio)",
                "desc": "Actualizar go-live. Validación rápida + validación ligera financieras.",
                "deliverable": "Transición cerrada para todas las empresas",
                "dur": 7,
            },
            {
                "code": "5.3",
                "name": "Estabilización final",
                "desc": "Revisión pipelines, optimización (particiones, Full/Incremental), documentación operativa.",
                "deliverable": "Proceso estable + documentación de operación",
                "dur": 7,
            },
        ],
    },
]

# Calcular fechas absolutas por actividad (secuencial dentro de cada fase, fases secuenciales)
def build_schedule(phases, start):
    phase_start = start
    for ph in phases:
        # La fase ocupa duration_days total; actividades se distribuyen dentro
        act_start = phase_start
        for act in ph["activities"]:
            act["start"] = act_start
            act["end"] = act_start + timedelta(days=act["dur"] - 1)
            act_start = act["end"] + timedelta(days=1)
        ph["start"] = phase_start
        ph["end"] = phase_start + timedelta(days=ph["duration_days"] - 1)
        phase_start = ph["end"] + timedelta(days=1)

build_schedule(PHASES, START_DATE)

# Total weeks in the project
total_end = PHASES[-1]["end"]
total_days = (total_end - START_DATE).days + 1
total_weeks = (total_days // 7) + (1 if total_days % 7 else 0)

# ─────────────────────────────────────────────
# 2. CREAR WORKBOOK
# ─────────────────────────────────────────────

wb = Workbook()
ws_gantt = wb.active
ws_gantt.title = "Gantt"
ws_detail = wb.create_sheet("Actividades Detalle")
ws_summary = wb.create_sheet("Resumen Fases")

# ─────────────────────────────────────────────
# HELPERS DE ESTILO
# ─────────────────────────────────────────────

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)

def bold_font(size=10, color="000000", italic=False):
    return Font(bold=True, size=size, color=color, italic=italic)

def normal_font(size=9, color="000000"):
    return Font(size=size, color=color)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left_wrap():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

WHITE = "FFFFFF"
HEADER_BG = "1F3864"   # azul oscuro header principal
SUBHEADER_BG = "2F5496"
GRAY_ROW = "F2F2F2"
DARK_GRAY = "595959"

# ─────────────────────────────────────────────
# 3. HOJA GANTT
# ─────────────────────────────────────────────

ws = ws_gantt
ws.sheet_view.showGridLines = False

# Columnas fijas
# A: Fase | B: Código | C: Actividad | D: Entregable | E: Días | F: Inicio | G: Fin | H: Responsable
# Luego semanas: col 9 en adelante

FIXED_COLS = 8
COL_FASE   = 1
COL_CODE   = 2
COL_ACT    = 3
COL_DELIV  = 4
COL_DAYS   = 5
COL_START  = 6
COL_END    = 7
COL_OWNER  = 8

# Anchos columnas fijas
ws.column_dimensions["A"].width = 20
ws.column_dimensions["B"].width = 7
ws.column_dimensions["C"].width = 38
ws.column_dimensions["D"].width = 32
ws.column_dimensions["E"].width = 7
ws.column_dimensions["F"].width = 11
ws.column_dimensions["G"].width = 11
ws.column_dimensions["H"].width = 16

# Anchos columnas de semana
for i in range(total_weeks):
    col_letter = get_column_letter(FIXED_COLS + 1 + i)
    ws.column_dimensions[col_letter].width = 4.5

# ---- TÍTULO ----
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=FIXED_COLS + total_weeks)
title_cell = ws.cell(1, 1, "PLAN DE ACCIÓN – INTEGRACIÓN SAP EN FABRIC  |  Arquitectura Medallion (Estrategia 3: Silver dual + Capa Conformada)")
title_cell.fill = fill(HEADER_BG)
title_cell.font = Font(bold=True, size=13, color=WHITE)
title_cell.alignment = center()
ws.row_dimensions[1].height = 28

# ---- SUBTÍTULO / meta ----
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=FIXED_COLS + total_weeks)
sub = ws.cell(2, 1, f"Inicio: {START_DATE.strftime('%d-%b-%Y')}   |   Duración total estimada: 13–20 semanas   |   Capacidad asignada: ~40–50% del equipo   |   Fecha de corte: 17-Mar-2026")
sub.fill = fill(SUBHEADER_BG)
sub.font = Font(size=9, color=WHITE, italic=True)
sub.alignment = center()
ws.row_dimensions[2].height = 18

# ---- ALERTA go-live SEL ----
ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=FIXED_COLS + total_weeks)
alert = ws.cell(3, 1, "⚠️  ALERTA: Go-live SEL objetivo Abril 2026.  Con duraciones mínimas y trabajo paralelo es alcanzable, pero requiere comenzar Fase 0 INMEDIATAMENTE y gestionar riesgos activamente.")
alert.fill = fill("FFF2CC")
alert.font = Font(bold=True, size=9, color="7F6000")
alert.alignment = center()
ws.row_dimensions[3].height = 18

# ---- FILA ENCABEZADOS COLUMNAS FIJAS ----
headers_row = 4
ws.row_dimensions[headers_row].height = 36
headers = ["Fase", "Cód.", "Actividad", "Entregable clave", "Días", "Inicio", "Fin", "Responsable"]
for c, h in enumerate(headers, 1):
    cell = ws.cell(headers_row, c, h)
    cell.fill = fill(DARK_GRAY)
    cell.font = bold_font(9, WHITE)
    cell.alignment = center()
    cell.border = thin_border()

# ---- ENCABEZADOS SEMANAS ----
week_monday = START_DATE - timedelta(days=START_DATE.weekday())  # lunes de la semana inicial
for i in range(total_weeks):
    col = FIXED_COLS + 1 + i
    wk_start = week_monday + timedelta(weeks=i)
    label = f"S{i+1}\n{wk_start.strftime('%d/%m')}"
    cell = ws.cell(headers_row, col, label)
    cell.fill = fill(DARK_GRAY)
    cell.font = bold_font(8, WHITE)
    cell.alignment = center()
    cell.border = thin_border()

# ---- FILAS DE ACTIVIDADES ----
current_row = headers_row + 1

for ph in PHASES:
    ph_num  = ph["num"]
    ph_color  = PHASE_COLORS[ph_num]
    ph_light  = PHASE_LIGHT[ph_num]

    # ---- fila de fase (título de bloque) ----
    ws.row_dimensions[current_row].height = 22
    ws.merge_cells(start_row=current_row, start_column=COL_FASE, end_row=current_row, end_column=COL_OWNER)
    phase_label = f"FASE {ph_num}: {ph['name'].upper()}  —  Objetivo: {ph['objetivo']}  |  Hito: {ph['hito']}"
    cell = ws.cell(current_row, COL_FASE, phase_label)
    cell.fill = fill(ph_color)
    cell.font = bold_font(9, WHITE)
    cell.alignment = left_wrap()

    # Barra de fase en el Gantt
    ph_start_week = (ph["start"] - START_DATE).days // 7
    ph_end_week   = (ph["end"]   - START_DATE).days // 7
    for w in range(total_weeks):
        col = FIXED_COLS + 1 + w
        cell = ws.cell(current_row, col, "")
        if ph_start_week <= w <= ph_end_week:
            cell.fill = fill(ph_color)
        else:
            cell.fill = fill("EEEEEE")
        cell.border = thin_border()

    current_row += 1

    # ---- filas de actividades ----
    for idx, act in enumerate(ph["activities"]):
        ws.row_dimensions[current_row].height = 30
        row_bg = ph_light if idx % 2 == 0 else WHITE

        # Columnas fijas
        data = [
            (COL_FASE,  f"  F{ph_num}",                                              9),
            (COL_CODE,  act["code"],                                                  9),
            (COL_ACT,   act["name"],                                                  9),
            (COL_DELIV, act["deliverable"],                                           8),
            (COL_DAYS,  act["dur"],                                                   9),
            (COL_START, act["start"].strftime("%d-%b-%y"),                            9),
            (COL_END,   act["end"].strftime("%d-%b-%y"),                              9),
            (COL_OWNER, "",                                                           9),
        ]
        for (col, val, sz) in data:
            cell = ws.cell(current_row, col, val)
            cell.fill = fill(row_bg)
            cell.font = normal_font(sz)
            cell.alignment = left_wrap() if col in (COL_ACT, COL_DELIV) else center()
            cell.border = thin_border()

        # Barra Gantt de actividad
        act_start_week = (act["start"] - START_DATE).days // 7
        act_end_week   = (act["end"]   - START_DATE).days // 7
        for w in range(total_weeks):
            col = FIXED_COLS + 1 + w
            cell = ws.cell(current_row, col, "")
            if act_start_week <= w <= act_end_week:
                cell.fill = fill(ph_color)
            else:
                cell.fill = fill(row_bg)
            cell.border = thin_border()

        current_row += 1

    # fila separadora
    ws.row_dimensions[current_row].height = 5
    for c in range(1, FIXED_COLS + total_weeks + 1):
        ws.cell(current_row, c, "").fill = fill("CCCCCC")
    current_row += 1

# ---- LEYENDA ----
current_row += 1
ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
ws.cell(current_row, 1, "LEYENDA").font = bold_font(10, HEADER_BG)
ws.cell(current_row, 1).alignment = left_wrap()
current_row += 1
for ph in PHASES:
    ws.cell(current_row, 1, f"  Fase {ph['num']}: {ph['name']}").fill = fill(PHASE_COLORS[ph['num']])
    ws.cell(current_row, 1).font = bold_font(9, WHITE)
    ws.cell(current_row, 1).alignment = left_wrap()
    current_row += 1

# Freeze panes
ws.freeze_panes = ws.cell(5, FIXED_COLS + 1)

# ─────────────────────────────────────────────
# 4. HOJA ACTIVIDADES DETALLE
# ─────────────────────────────────────────────

ws2 = ws_detail
ws2.sheet_view.showGridLines = False

# Título
ws2.merge_cells("A1:G1")
c = ws2.cell(1, 1, "ACTIVIDADES DETALLE – Integración SAP en Fabric")
c.fill = fill(HEADER_BG); c.font = bold_font(12, WHITE); c.alignment = center()
ws2.row_dimensions[1].height = 26

# Headers
headers2 = ["Fase", "Código", "Actividad", "Descripción detallada", "Entregable", "Días", "Fechas"]
widths2   = [20,    8,        35,          55,                       38,           7,      22]
for ci, (h, w) in enumerate(zip(headers2, widths2), 1):
    ws2.column_dimensions[get_column_letter(ci)].width = w
    cell = ws2.cell(2, ci, h)
    cell.fill = fill(DARK_GRAY); cell.font = bold_font(9, WHITE)
    cell.alignment = center(); cell.border = thin_border()
ws2.row_dimensions[2].height = 20

row2 = 3
for ph in PHASES:
    # Cabecera de fase
    ws2.merge_cells(start_row=row2, start_column=1, end_row=row2, end_column=7)
    ph_label = f"FASE {ph['num']}: {ph['name'].upper()}   |   Objetivo: {ph['objetivo']}   |   Hito: {ph['hito']}   |   {ph['start'].strftime('%d-%b-%y')} → {ph['end'].strftime('%d-%b-%y')}"
    c = ws2.cell(row2, 1, ph_label)
    c.fill = fill(PHASE_COLORS[ph['num']]); c.font = bold_font(9, WHITE); c.alignment = left_wrap()
    ws2.row_dimensions[row2].height = 20
    row2 += 1

    for idx, act in enumerate(ph["activities"]):
        ws2.row_dimensions[row2].height = 44
        bg = PHASE_LIGHT[ph['num']] if idx % 2 == 0 else WHITE
        vals = [
            f"F{ph['num']} – {ph['name']}",
            act["code"],
            act["name"],
            act["desc"],
            act["deliverable"],
            act["dur"],
            f"{act['start'].strftime('%d-%b-%y')} → {act['end'].strftime('%d-%b-%y')}",
        ]
        for ci, v in enumerate(vals, 1):
            cell = ws2.cell(row2, ci, v)
            cell.fill = fill(bg); cell.font = normal_font(9)
            cell.alignment = left_wrap() if ci in (3, 4, 5) else center()
            cell.border = thin_border()
        row2 += 1

    row2 += 1  # espacio entre fases

# ─────────────────────────────────────────────
# 5. HOJA RESUMEN FASES
# ─────────────────────────────────────────────

ws3 = ws_summary
ws3.sheet_view.showGridLines = False

ws3.merge_cells("A1:H1")
c = ws3.cell(1, 1, "RESUMEN EJECUTIVO – Plan de Acción Integración SAP en Fabric")
c.fill = fill(HEADER_BG); c.font = bold_font(13, WHITE); c.alignment = center()
ws3.row_dimensions[1].height = 28

ws3.merge_cells("A2:H2")
c = ws3.cell(2, 1, "Estrategia 3: Silver dual + Capa Conformada  |  Duración total estimada: 13–20 semanas  |  Capacidad: ~40–50% del equipo")
c.fill = fill(SUBHEADER_BG); c.font = Font(size=9, color=WHITE, italic=True); c.alignment = center()
ws3.row_dimensions[2].height = 16

headers3 = ["Fase", "Nombre", "Objetivo", "Hito de cierre", "Duración", "Inicio", "Fin", "N° Actividades"]
widths3   = [7,      22,       55,         35,                14,         12,       12,    16]
for ci, (h, w) in enumerate(zip(headers3, widths3), 1):
    ws3.column_dimensions[get_column_letter(ci)].width = w
    cell = ws3.cell(4, ci, h)
    cell.fill = fill(DARK_GRAY); cell.font = bold_font(9, WHITE)
    cell.alignment = center(); cell.border = thin_border()
ws3.row_dimensions[4].height = 20

for idx, ph in enumerate(PHASES):
    row3 = 5 + idx
    ws3.row_dimensions[row3].height = 36
    bg = PHASE_LIGHT[ph['num']]
    vals = [
        ph['num'],
        ph['name'],
        ph['objetivo'],
        ph['hito'],
        f"{ph['duration_days']} días",
        ph['start'].strftime('%d-%b-%y'),
        ph['end'].strftime('%d-%b-%y'),
        len(ph['activities']),
    ]
    for ci, v in enumerate(vals, 1):
        cell = ws3.cell(row3, ci, v)
        cell.fill = fill(PHASE_COLORS[ph['num']] if ci == 1 else bg)
        cell.font = bold_font(9, WHITE) if ci == 1 else normal_font(9)
        cell.alignment = center() if ci not in (3, 4) else left_wrap()
        cell.border = thin_border()

# Resumen total
row3 = 5 + len(PHASES) + 1
ws3.merge_cells(start_row=row3, start_column=1, end_row=row3, end_column=4)
c = ws3.cell(row3, 1, "TOTAL ESTIMADO (escenario optimista, mínimos, sin buffers):")
c.fill = fill(HEADER_BG); c.font = bold_font(10, WHITE); c.alignment = left_wrap()
c = ws3.cell(row3, 5, f"{sum(p['duration_days'] for p in PHASES)} días")
c.fill = fill(HEADER_BG); c.font = bold_font(10, WHITE); c.alignment = center()
ws3.cell(row3, 6, START_DATE.strftime('%d-%b-%y')).fill = fill(HEADER_BG)
ws3.cell(row3, 6).font = bold_font(10, WHITE); ws3.cell(row3, 6).alignment = center()
ws3.cell(row3, 7, total_end.strftime('%d-%b-%y')).fill = fill(HEADER_BG)
ws3.cell(row3, 7).font = bold_font(10, WHITE); ws3.cell(row3, 7).alignment = center()
ws3.cell(row3, 8, sum(len(p['activities']) for p in PHASES)).fill = fill(HEADER_BG)
ws3.cell(row3, 8).font = bold_font(10, WHITE); ws3.cell(row3, 8).alignment = center()
ws3.row_dimensions[row3].height = 22

row3 += 2
ws3.merge_cells(start_row=row3, start_column=1, end_row=row3, end_column=8)
note = ("NOTA: Las estimaciones asumen ejecución en paralelo a operación habitual (Fabric, RPAs, Sistema de Indicadores). "
        "Con ~40–50% de capacidad asignada a SAP. Si se reserva más capacidad en periodos críticos (pre go-live SEL), "
        "las fases se pueden acortar. El escenario realista es 15–17 semanas.")
c = ws3.cell(row3, 1, note)
c.fill = fill("FFF2CC"); c.font = Font(size=9, color="7F6000", italic=True); c.alignment = left_wrap()
ws3.row_dimensions[row3].height = 46

# ─────────────────────────────────────────────
# 6. GUARDAR
# ─────────────────────────────────────────────

output_path = "/Users/octavio.cruz/Desktop/Fabric SAP/Plan_Accion_SAP_Fabric_Gantt.xlsx"
wb.save(output_path)
print(f"\n✅  Archivo generado: {output_path}")
print(f"    Fases:       {len(PHASES)}")
print(f"    Actividades: {sum(len(p['activities']) for p in PHASES)}")
print(f"    Semanas Gantt: {total_weeks}")
print(f"    Inicio: {START_DATE}  |  Fin: {total_end}")
